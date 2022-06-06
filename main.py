import openpyxl
from openpyxl import Workbook
wb = Workbook()#openpyxl.load_workbook('sample.xlsx')
print(type(wb))
print(wb.sheetnames)
for sheet in wb:
     print(sheet.title)
ws = wb.active
ws['A1']="ciao"
ws['B1']="ciao mondo"
wb.save('prova.xlsx')
#print("Ciao Mondo!!!")
#https://openpyxl.readthedocs.io/en/stable/tutorial.html
